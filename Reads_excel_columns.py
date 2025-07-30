import pandas as pd
import argparse 
import os
import numpy as np
import openpyxl

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('month', nargs='?', default='January', help='Monat für Filterung') 
    parser.add_argument('--recl', default='recl.xlsx', help='Pfad zur recl.xlsx Datei')
    parser.add_argument('--grp', default='grp.xlsx', help='Pfad zur grp.xlsx Datei')
    return parser.parse_args()

# Parcer ergänzt
args = parse_args()

print(f"Verarbeitung für Monat: {args.month}")
print(f"Recl-Datei: {args.recl}")

# Überprüfen ob die Datei exists
if not os.path.exists(args.recl):
    print(f"Fehler: Datei {args.recl} nicht gefunden!")
    exit(1)

# 1. Rohdaten komplett einlesen (ohne Header)
df_raw = pd.read_excel(args.recl, header=None, engine='openpyxl')
df_grp = pd.read_excel(args.grp, header=None, engine='openpyxl') 
df_raw.to_excel('file2_raw.xlsx', header=False, index=False)
print("Rohdaten: file2_raw.xlsx")

# 2. Erste 3 Zeilen entfernen
df_no_rows = df_raw.iloc[3:].reset_index(drop=True)

# 3. Spalten entfernen:
#    - die ersten 4 Spalten (0–3)
#    - zusätzlich Spalten 11, 14, 15, 16, 17, 21, 22, 23, 24, 25
#    - sowie Spalten 27 bis 41 (inklusive)
cols_to_drop = [0, 1, 2, 3,
                11, 14, 15, 17,
                21, 22, 23, 24, 25] + list(range(27, 41))

df_processed = df_no_rows.drop(columns=cols_to_drop, errors='ignore')
df_processed.to_excel('file2_processed.xlsx', header=False, index=False)
print("Ohne ausgewählte Spalten: file2_processed.xlsx")

# 4. Filtern:
#    – Die erste Zeile (Header) unberührt lassen
header_row = df_processed.iloc[[0]]
data_rows  = df_processed.iloc[1:]

#    – Nur Zeilen, in denen Spalte 18 == "Einsteller"
#      UND Spalte 7 mit "AG" beginnt
mask = (
    (data_rows[18] == 'Einsteller') &
    (data_rows[7].astype(str).str.startswith('AG')) &
    (data_rows[19] != 'Wurde abgelehnt'))

filtered_rows = data_rows[mask]

# Monatfilterung ergänzt
if args.month and len(filtered_rows) > 0:
    # Monatszuordnung
    month_map = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }

    try:
        print(f"\n=== Monatsfilterung für {args.month} ===")
        
        # Nach dem Löschen der Spalten ist die Datumsspalte jetzt Spalte 0
        # (weil Spalten 0,1,2,3 gelöscht wurden)

        date_column_position = 0  # Erste Spalte (positionsmäßig)
        print(f"Verwende Spalte an Position {date_column_position} für Datumsfilterung")

        print(f"Erste 5 Werte in Datumsspalte:")
        print(filtered_rows.iloc[:5, date_column_position] if len(filtered_rows) > 0 else "Keine Daten")
        
        # Datumskonvertierung mit dem richtigen Format DD/MM/YYYY
        date_series = pd.to_datetime(
            filtered_rows.iloc[:, 0],  
            errors='coerce'
        )
        
        # Wenn nicht alle Daten erkannt wurden, wir versuchen spezifische Formate
        if date_series.notna().sum() < len(filtered_rows) * 0.5:  # Wenn weniger als 50% erkannt
            for fmt in date_formats:
                try:
                    date_series = pd.to_datetime(filtered_rows[0], format=fmt, errors='coerce')
                    if date_series.notna().sum() > 0:
                        print(f"Verwendetes Datumsformat: {fmt}")
                        break
                except:
                    continue
        
        # Debug-Informationen
        valid_dates = date_series.notna().sum()
        print(f"Erfolgreich konvertierte Datumsangaben: {valid_dates} von {len(date_series)}")
        
        month_number = month_map.get(args.month, 1)
        print(f"Filtern nach Monat: {args.month} (Nummer {month_number})")

        if valid_dates > 0:
            available_months = sorted(date_series.dt.month.dropna().unique())
            print(f"Verfügbare Monate in den Daten: {available_months}")
            

            # Prüfen ob der gewünschte Monat in den Daten vorhanden ist
            if month_number not in available_months:
                print(f"ACHTUNG: Monat {args.month} ({month_number}) ist in den Daten nicht vorhanden!")
            
            # Monatsfilterung anwenden
            month_mask = date_series.dt.month == month_number
            rows_before = len(filtered_rows)
            filtered_rows = filtered_rows[month_mask]
            rows_after = len(filtered_rows)
            
            print(f"Zeilen vor Filter: {rows_before}")
            print(f"Zeilen nach Filter: {rows_after}")
            
            print(f"Gefiltert nach Monat: {args.month}")
        else:
            print("Keine Datumsangaben konnten konvertiert werden!")
            
    except Exception as e:
        print(f"Fehler bei der Monatsfilterung: {e}")
        import traceback
        traceback.print_exc()

#    Header und gefilterte Daten wieder zusammenfügen
df_final = pd.concat([header_row, filtered_rows], ignore_index=True)

# 5. Timestamp-Spalten 0 und 6 in reines Datum wandeln (ab Zeile 1)
for col_idx in (0, 6):
    df_final.iloc[1:, col_idx] = pd.to_datetime(
        df_final.iloc[1:, col_idx],
        format="%d.%m.%Y %H:%M:%S", 
        # infer_datetime_format=True,
        errors='coerce'
    ).dt.date

# 6. ERGEBNIS IN EINER EXCEL-DATEI MIT MEHREREN REGISTERKARTEN SPEICHERN
result_filename = f'Ergebnis_{args.month}.xlsx'

try:
    # Erste Registerkarte: "Alle" - die gefilterten Daten
    with pd.ExcelWriter(result_filename, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Alle', index=False, header=False)
    
        # Greifen auf das Worksheet-Objekt zu
        worksheet_alle = writer.sheets['Alle']
        # Setzen die Breite auf das 3-fache der Standardbreite
        # Standardbreite ist oft ~8.43, wir setzen z.B. 25
        worksheet_alle.column_dimensions['A'].width = 15
        worksheet_alle.column_dimensions['B'].width = 10
        worksheet_alle.column_dimensions['C'].width = 15
        worksheet_alle.column_dimensions['D'].width = 10
        worksheet_alle.column_dimensions['E'].width = 10
        worksheet_alle.column_dimensions['F'].width = 10
        worksheet_alle.column_dimensions['G'].width = 15
        worksheet_alle.column_dimensions['H'].width = 25
        worksheet_alle.column_dimensions['I'].width = 25
        worksheet_alle.column_dimensions['J'].width = 40
        worksheet_alle.column_dimensions['K'].width = 10
        worksheet_alle.column_dimensions['L'].width = 20
        worksheet_alle.column_dimensions['M'].width = 20
        worksheet_alle.column_dimensions['N'].width = 30

    print(f"Gefilterte Daten gespeichert in Registerkarte 'Alle' der Datei: {result_filename}")
    
except Exception as e:
    print(f"Fehler beim Speichern der gefilterten Daten: {e}")
    import traceback
    traceback.print_exc()

# 7. Erledigt und Offen
try:
    print(f"\n=== Erledigt und Offen Filterung ===")
    
    if len(df_final) > 1 and len(df_final.columns) > 5:
        # Header und Daten trennen
        header = df_final.iloc[0:1]  # Erste Zeile (Header)
        data = df_final.iloc[1:]     # Datenzeilen
        
        # Spalte 5 enthält den Status (Erledigt/Offen)
        status_column = 5
        
        print(f"Verfügbare Status in Spalte {status_column}:")
        print(data.iloc[:, status_column].value_counts(dropna=False))
        
        # Filter für "Erledigt"
        erledigt_mask = data.iloc[:, status_column].astype(str).str.strip() == 'erledigt'
        erledigt_data = data[erledigt_mask]
        print(f"Anzahl 'Erledigt': {len(erledigt_data)}")
        
        # Filter für "Offen"
        offen_mask = data.iloc[:, status_column].astype(str).str.strip() == 'offen'
        offen_data = data[offen_mask]
        print(f"Anzahl 'Offen': {len(offen_data)}")
        
        # Daten mit Header kombinieren
        erledigt_final = pd.concat([header, erledigt_data], ignore_index=True)
        offen_final = pd.concat([header, offen_data], ignore_index=True)

        # Registerkarten hinzufügen (zweite und dritte Position)
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a') as writer:
            erledigt_final.to_excel(writer, sheet_name='Erledigt', index=False, header=False)
            offen_final.to_excel(writer, sheet_name='Offen', index=False, header=False)

            # Greifen auf das Worksheet-Objekt zu
            worksheet_erledigt = writer.sheets['Erledigt']

            worksheet_erledigt.column_dimensions['A'].width = 15
            worksheet_erledigt.column_dimensions['B'].width = 10
            worksheet_erledigt.column_dimensions['C'].width = 15
            worksheet_erledigt.column_dimensions['D'].width = 10
            worksheet_erledigt.column_dimensions['E'].width = 10
            worksheet_erledigt.column_dimensions['F'].width = 10
            worksheet_erledigt.column_dimensions['G'].width = 15
            worksheet_erledigt.column_dimensions['H'].width = 25
            worksheet_erledigt.column_dimensions['I'].width = 25
            worksheet_erledigt.column_dimensions['J'].width = 40
            worksheet_erledigt.column_dimensions['K'].width = 10
            worksheet_erledigt.column_dimensions['L'].width = 20
            worksheet_erledigt.column_dimensions['M'].width = 20
            worksheet_erledigt.column_dimensions['N'].width = 30


            # Greifen auf das Worksheet-Objekt zu
            worksheet_offen = writer.sheets['Offen']
            
            worksheet_offen.column_dimensions['A'].width = 15
            worksheet_offen.column_dimensions['B'].width = 10
            worksheet_offen.column_dimensions['C'].width = 15
            worksheet_offen.column_dimensions['D'].width = 10
            worksheet_offen.column_dimensions['E'].width = 10
            worksheet_offen.column_dimensions['F'].width = 10
            worksheet_offen.column_dimensions['G'].width = 15
            worksheet_offen.column_dimensions['H'].width = 25
            worksheet_offen.column_dimensions['I'].width = 25
            worksheet_offen.column_dimensions['J'].width = 40
            worksheet_offen.column_dimensions['K'].width = 10
            worksheet_offen.column_dimensions['L'].width = 20
            worksheet_offen.column_dimensions['M'].width = 20
            worksheet_offen.column_dimensions['N'].width = 30
        
        print("Registerkarten 'Erledigt' und 'Offen' hinzugefügt")
        
    else:
        print("Nicht genügend Daten oder Spalten für Status-Filterung")
        
except Exception as e:
    print(f"Fehler bei der Status-Filterung: {e}")
    import traceback
    traceback.print_exc()

# 8. Hauptthema Gruppierung und Analyse - IN ZWEITER REGISTERKARTE
try:
    print(f"\n=== Hauptthema Analyse für {args.month} ===")
    
    # Sicherstellen, dass genügend Daten und Spalten vorhanden sind
    if len(df_final) > 1 and len(df_final.columns) > 8:
        # Datenzeilen (ohne Header)
        data = df_final.iloc[1:]     
        
        print(f"Daten für Analyse: {len(data)} Zeilen")
        print(f"Verfügbare Spalten: {list(data.columns)}")

        # Debug: Zeige erste Werte der Hauptthema-Spalte
        print(f"Erste 5 Werte in Hauptthema-Spalte (Index 8):")
        print(data.iloc[:5, 8] if len(data) > 0 else "Keine Daten")
        
        # Gruppierung nach Hauptthema (Spalte 8)
        hauptthema_column = 8
        
        # Ersetze leere oder NaN Werte durch "Sonstiges"
        data_filled = data.copy()
        data_filled.iloc[:, hauptthema_column] = data_filled.iloc[:, hauptthema_column].fillna('Sonstiges')
        data_filled.iloc[:, hauptthema_column] = data_filled.iloc[:, hauptthema_column].replace('', 'Sonstiges')
        
        # Entferne komplett leere Zeilen (wenn alle Spalten leer sind)
        valid_data = data_filled[data_filled.iloc[:, hauptthema_column].notna()]

        if len(valid_data) > 0:
            # Gruppieren und zählen
            hauptthema_counts = valid_data.iloc[:, hauptthema_column].value_counts().reset_index()
            hauptthema_counts.columns = ['Hauptthema', 'Summe']
            
            # Prozentuale Anteile berechnen
            total_rows = len(valid_data)
            hauptthema_counts['In % gegenüber allen Beanstandungen'] = (
                hauptthema_counts['Summe'] / total_rows * 100
            ).round(2)
            
            # Ergebnis sortieren nach Anzahl (absteigend)
            hauptthema_counts = hauptthema_counts.sort_values('Summe', ascending=False)
            
            # GESAMT-Zeile am Ende hinzufügen
            gesamt_row = pd.DataFrame({
                'Hauptthema': ['Gesamt'],
                'Summe': [total_rows],
                'In % gegenüber allen Beanstandungen': [100.00]
            })
            
            # Kombiniere die Ergebnisse mit der Gesamtzeile
            hauptthema_analysis = pd.concat([hauptthema_counts, gesamt_row], ignore_index=True)
            
            print(f"Hauptthema Analyse Ergebnis:")
            print(hauptthema_analysis.head(15))

            # Registerkarte: "Hauptthema Analyse Ergebnis" hinzufügen
            with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a') as writer:
                hauptthema_analysis.to_excel(writer, sheet_name='Hauptthema Analyse Ergebnis', index=False)

                from openpyxl.chart import PieChart, Reference
                from openpyxl.chart.legend import Legend
                from openpyxl.chart.layout import Layout, ManualLayout
                from openpyxl.chart.label import DataLabelList
                from openpyxl.chart.series import DataPoint
                from openpyxl.drawing.fill import PatternFillProperties
                from openpyxl.drawing.colors import ColorChoice

                
                # Greifen auf das Worksheet-Objekt zu
                worksheet_haupthema = writer.sheets['Hauptthema Analyse Ergebnis']
                
                worksheet_haupthema.column_dimensions['A'].width = 25
                worksheet_haupthema.column_dimensions['B'].width = 15
                worksheet_haupthema.column_dimensions['C'].width = 40

                # Anzahl der Datenzeilen (ohne Gesamt-Zeile)
                n = len(hauptthema_analysis) - 1

                # 1) Kategorien (Hauptthemen) aus A2:A(n+1)
                labels = Reference(
                    worksheet_haupthema,
                    min_col=1,
                    min_row=2,
                    max_row=1 + n
                )

                # 2) Werte (Summe) aus B2:B(n+1)
                data = Reference(
                    worksheet_haupthema,
                    min_col=2,
                    min_row=2,
                    max_row=1 + n
                )

                # PieChart erstellen und konfigurieren
                pie = PieChart()
                pie.title = "Verteilung der Hauptthemen"
                pie.add_data(data, titles_from_data=False)
                pie.set_categories(labels)
                pie.style = 10 

                pie.legend = Legend()
                pie.legend.position = 'r'    
                pie.legend.overlay = False 

                pie.layout = Layout(
                    manualLayout=ManualLayout(
                        x=0.1, 
                        y=0.1,   
                        w=0.8,  
                        h=0.8    
                    )
                )

                pie.series[0].title = None

                pie.dataLabels = DataLabelList()
                pie.dataLabels.showSerName   = False
                pie.dataLabels.showPercent   = True
                pie.dataLabels.numFmt         = '0.00%'
                pie.dataLabels.showLegendKey = False
                pie.dataLabels.showCatName   = True

                # Diagrammgröße anpassen (Einheiten sind Excel-intern, ~1 = 1 Zoll)
                pie.width = 30   # Breite in Zoll
                pie.height = 15  # Höhe in Zoll

                # Diagramm in das Sheet einfügen (Position E5)
                worksheet_haupthema.add_chart(pie, "E2")

            print(f"Hauptthema Analyse gespeichert in Registerkarte 'Hauptthema Analyse Ergebnis'")
            print(f"Gesamtanzahl Beanstandungen: {total_rows}")
            print(f"Anzahl verschiedener Hauptthemen (ohne Gesamt): {len(hauptthema_analysis) - 1}")
        else:
            print("Keine gültigen Daten für Hauptthema Analyse gefunden")
    else:
        print("Nicht genügend Daten oder Spalten für Hauptthema Analyse")
        print(f"Verfügbare Zeilen: {len(df_final)}, Verfügbare Spalten: {len(df_final.columns) if len(df_final) > 0 else 0}")

except Exception as e:
    print(f"Fehler bei der Hauptthema Analyse: {e}")
    import traceback
    traceback.print_exc()

# 9. Pivot Einsteller Hauptthema
try:
    print(f"\n=== Pivot Einsteller Hauptthema für {args.month} ===")
    
    # Sicherstellen, dass genügend Daten und Spalten vorhanden sind
    # Benötigt: Hauptthema (Index 8) und Einsteller (Index 3)
    if len(df_final) > 1 and len(df_final.columns) > max(8, 3):
        # Datenzeilen (ohne Header)
        data = df_final.iloc[1:]
        
        print(f"Daten für Pivot-Analyse: {len(data)} Zeilen")
        
        # Spaltenindizes
        einsteller_column = 3  # Spalte Einsteller
        hauptthema_column = 8  # Spalte Hauptthema
        
        # Debug: Zeige einige Werte
        # print(f"Erste 5 Werte in Einsteller-Spalte (Index {einsteller_column}):")
        # print(data.iloc[:5, einsteller_column] if len(data) > 0 else "Keine Daten")
        # print(f"Erste 5 Werte in Hauptthema-Spalte (Index {hauptthema_column}):")
        # print(data.iloc[:5, hauptthema_column] if len(data) > 0 else "Keine Daten")
        
        # Entferne Zeilen mit fehlenden Werten in den Schlüsselspalten
        # und Zeilen mit leerem String in diesen Spalten
        valid_pivot_data = data.dropna(subset=[data.columns[einsteller_column], data.columns[hauptthema_column]])
        valid_pivot_data = valid_pivot_data.copy() # Um SettingWithCopyWarning zu vermeiden
        valid_pivot_data.loc[:, 'Einsteller_Clean'] = valid_pivot_data.iloc[:, einsteller_column].astype(str).str.strip()
        valid_pivot_data.loc[:, 'Hauptthema_Clean'] = valid_pivot_data.iloc[:, hauptthema_column].astype(str).str.strip()
        
        valid_pivot_data = valid_pivot_data[
            (valid_pivot_data['Einsteller_Clean'] != '') &
            (valid_pivot_data['Hauptthema_Clean'] != '')
        ]
        
        print(f"Gültige Daten für Pivot: {len(valid_pivot_data)} Zeilen")
        
        if len(valid_pivot_data) > 0:
            # Erstelle die Pivot-Tabelle
            # Werte: Anzahl der Zeilen (size)
            # Index: Hauptthema
            # Columns: Einsteller
            # fill_value=0 sorgt dafür, dass leere Zellen 0 enthalten
            pivot_table = pd.pivot_table(
                valid_pivot_data,
                values=data.columns[0], # Wir nehmen eine beliebige Spalte, da wir 'size' als aggfunc verwenden
                index='Hauptthema_Clean',  # Hauptthema
                columns='Einsteller_Clean', # Einsteller
                aggfunc='size', # Zähle die Anzahl der Zeilen
                fill_value=0,
                sort=False # Reihenfolge wie im Original
            )
            
            # Sortiere die Zeilen nach Gesamtanzahl (absteigend)
            # Berechne die Summe pro Zeile *bevor* die Gesamtzeile hinzugefügt wird
            row_totals = pivot_table.sum(axis=1)
            
            # Füge die Gesamt-Spalte hinzu
            pivot_table['Gesamt'] = row_totals
            
            # Sortiere die Tabelle nach der Gesamt-Spalte (absteigend)
            pivot_table_sorted = pivot_table.sort_values(by='Gesamt', ascending=False)
            
            # Füge eine Summenzeile am Ende hinzu
            column_totals = pivot_table_sorted.sum(axis=0)
            column_totals.name = 'Gesamt'
            pivot_table_final = pd.concat([pivot_table_sorted, column_totals.to_frame().T])
            
            print("Pivot-Tabelle (Ausschnitt):")
            print(pivot_table_final.head(10))
            
            # Letzte Registerkarte: "Pivot Einsteller Hauptthema" hinzufügen
            with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                pivot_table_final.to_excel(writer, sheet_name='Pivot Einsteller Hauptthema')

                # Greifen auf das Worksheet-Objekt zu
                worksheet_haupthema = writer.sheets['Pivot Einsteller Hauptthema']
                
                worksheet_haupthema.column_dimensions['A'].width = 25
            
            print(f"Pivot-Tabelle gespeichert in Registerkarte 'Pivot Einsteller Hauptthema'")
            print(f"Anzahl der Hauptthemen (exkl. Gesamtzeile): {len(pivot_table_sorted)}")
            print(f"Anzahl der Einsteller (exkl. Gesamtspalte): {len(pivot_table_final.columns) - 1}") # -1 für 'Gesamt'-Spalte
            
        else:
            print("Keine gültigen Daten für Pivot-Erstellung gefunden")
            # Erstelle eine leere Tabelle mit passendem Namen
            empty_pivot = pd.DataFrame({'Hinweis': ['Keine Daten für Pivot-Tabelle verfügbar']})
            with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                empty_pivot.to_excel(writer, sheet_name='Pivot Einsteller Hauptthema', index=False)
            
    else:
        print("Nicht genügend Daten oder Spalten für Pivot-Erstellung")
        print(f"Verfügbare Zeilen: {len(df_final)}, Verfügbare Spalten: {len(df_final.columns) if len(df_final) > 0 else 0}")
        # Erstelle eine leere Tabelle mit passendem Namen
        empty_pivot = pd.DataFrame({'Hinweis': ['Nicht genügend Spalten für Pivot-Tabelle']})
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            empty_pivot.to_excel(writer, sheet_name='Pivot Einsteller Hauptthema', index=False)

except Exception as e:
    print(f"Fehler bei der Pivot-Erstellung: {e}")
    import traceback
    traceback.print_exc()

# 10. Spalten entfernen Gruppenreporting:
cols_to_drop_grp = [0, 1,
                6, 7, 9, 10, 11,
                12, 13, 14, 16, 17] + list(range(18, 31))

df_processed_grp = df_grp.drop(columns=cols_to_drop_grp, errors='ignore')
df_processed_grp.to_excel('grp_processed.xlsx', header=False, index=False)
print(f"Verfügbare Spalten in df_processed_grp nach Löschen: {list(df_processed_grp.columns)}")
print("Ohne ausgewählte Spalten: grp_processed.xlsx")

# 11. Filtern:
#    – Die erste Zeile (Header) unberührt lassen
header_row_grp = df_processed_grp.iloc[[0]]
data_rows_grp  = df_processed_grp.iloc[1:]

# Debug: Проверьте структуру данных
print(f"Verfügbare Spalten in df_processed_grp: {list(df_processed_grp.columns)}")
print(f"Erste Zeile der Daten:")
print(data_rows_grp.iloc[0] if len(data_rows_grp) > 0 else "Keine Daten")

#    – Filtern nach Bedingungen (passen Sie die Spaltenpositionen an)
# Beispiel: Angenommen, nach dem Löschen ist:
# Spalte 0: 'Verkauft'/anderer Wert
# Spalte 1: 'AG...'
if len(data_rows_grp.columns) > 1:
    mask_grp = (
        (data_rows_grp.iloc[:, 0] == 'Verkauft') &
        (data_rows_grp.iloc[:, 1].astype(str).str.startswith('AG'))
    )
    filtered_rows_grp = data_rows_grp[mask_grp]
    print(f"Nach Filter: {len(filtered_rows_grp)} Zeilen")
else:
    print("Nicht genügend Spalten für GRP-Filter")
    filtered_rows_grp = data_rows_grp.iloc[0:0]

# 12. Monatfilterung ergänzt
if args.month and len(filtered_rows_grp) > 0 and len(filtered_rows_grp.columns) > 4:
    # Monatszuordnung
    month_map = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }

    try:
        print(f"\n=== Monatsfilterung für {args.month} (GRP) ===")
        
        # Datumsspalte position bestimmen
        date_column_position = 4  
        print(f"Verwende Spalte an Position {date_column_position} für Datumsfilterung")

        if len(filtered_rows_grp.columns) > date_column_position:
            print(f"Erste 5 Werte in Datumsspalte:")
            print(filtered_rows_grp.iloc[:4, date_column_position] if len(filtered_rows_grp) > 0 else "Keine Daten")
            
            # Datumskonvertierung
            date_series = pd.to_datetime(
                filtered_rows_grp.iloc[:, date_column_position], 
                errors='coerce'
            )
            
            # Debug-Informationen
            valid_dates_grp = date_series.notna().sum()
            print(f"Erfolgreich konvertierte Datumsangaben: {valid_dates_grp} von {len(date_series)}")
            
            month_number = month_map.get(args.month, 1)
            print(f"Filtern nach Monat: {args.month} (Nummer {month_number})")

            if valid_dates_grp > 0:
                available_months = sorted([int(m) for m in date_series.dt.month.dropna().unique()])
                print(f"Verfügbare Monate in den Daten: {available_months}")
                
                # Prüfen ob der gewünschte Monat in den Daten vorhanden ist
                if month_number not in available_months:
                    print(f"ACHTUNG: Monat {args.month} ({month_number}) ist in den Daten nicht vorhanden!")
                
                # Monatsfilterung anwenden
                month_mask = date_series.dt.month == month_number
                rows_before = len(filtered_rows_grp)
                filtered_rows_grp = filtered_rows_grp[month_mask]  
                rows_after = len(filtered_rows_grp)  
                
                print(f"Zeilen vor Filter: {rows_before}")
                print(f"Zeilen nach Filter: {rows_after}")
                
                print(f"Gefiltert nach Monat: {args.month}")
            else:
                print("Keine Datumsangaben konnten konvertiert werden!")
        else:
            print(f"Nicht genügend Spalten: Benötige mindestens {date_column_position + 1} Spalten")
            
    except Exception as e:
        print(f"Fehler bei der Monatsfilterung: {e}")
        import traceback
        traceback.print_exc()

#    Header und gefilterte Daten wieder zusammenfügen
df_final_grp = pd.concat([header_row_grp, filtered_rows_grp], ignore_index=True) 

# 13. Ergebnis Gruppenreporting speichern
try:
    # Hinzufügen zum bestehenden Excel-File
    with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_final_grp.to_excel(writer, sheet_name='Gruppenreporting', index=False, header=False)

        # Greifen auf das Worksheet-Objekt zu
        worksheet_gruppenreporting = writer.sheets['Gruppenreporting']
        # Standardbreite ist oft ~8.43, wir setzen z.B. 25
        worksheet_gruppenreporting.column_dimensions['A'].width = 15
        worksheet_gruppenreporting.column_dimensions['B'].width = 15
        worksheet_gruppenreporting.column_dimensions['C'].width = 35
        worksheet_gruppenreporting.column_dimensions['D'].width = 15
        worksheet_gruppenreporting.column_dimensions['E'].width = 35
        worksheet_gruppenreporting.column_dimensions['F'].width = 25

    print(f"Gefilterte Daten gespeichert in Registerkarte 'Gruppenreporting' der Datei: {result_filename}")
    
except Exception as e:
    print(f"Fehler beim Speichern der gefilterten Daten: {e}")
    import traceback
    traceback.print_exc()

# 14. Verkaufsstatistik nach User erstellen Gruppenreporting
try:
    print(f"\n=== Verkaufsstatistik nach User ===")
    
    # Sicherstellen, dass genügend Daten und Spalten vorhanden sind
    # Wir brauchen mindestens 2 Zeilen (Header + Daten) und 5 Spalten (für Spalte 4)
    if len(df_final_grp) > 1 and len(df_final_grp.columns) > 1:
        # Header trennen und Datenzeilen erhalten
        # data_rows_grp_final = df_final_grp.iloc[1:] # Wenn df_final_grp der gefilterte DataFrame ist
        
        # Da wir direkt aus filtered_rows_grp arbeiten (vor dem Zusammenfügen mit Header):
        if 'filtered_rows_grp' in locals() and len(filtered_rows_grp) > 0:
            data_for_sales = filtered_rows_grp
        else:
            print("Keine Daten für Verkaufsstatistik verfügbar.")
            # Leere Statistik erstellen
            empty_sales = pd.DataFrame({'Hinweis': ['Keine Verkaufsdaten verfügbar']})
            with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                empty_sales.to_excel(writer, sheet_name='Verkäufe nach User', index=False, header=False)
            print("Leere Registerkarte 'Verkäufe nach User' hinzugefügt.")
            raise Exception("Keine Daten für Verkaufsstatistik")
            
        # User-Spalte ist Spalte 4 (0-basiert Index 4)
        user_column_index = 1
        
        print(f"Daten für Verkaufsstatistik: {len(data_for_sales)} Zeilen")
        print(f"User-Spalte (Index {user_column_index}):")
        print(data_for_sales.iloc[:3, user_column_index] if len(data_for_sales) > 0 else "Keine Daten")
        
        # Gruppieren nach User (Spalte 4) und zählen
        # value_counts() zählt automatisch die Anzahl der Zeilen pro eindeutigem Wert
        # dropna=False: Zähle auch NaN-Werte als separate Kategorie (kann angepasst werden)
        user_counts = data_for_sales.iloc[:, user_column_index].value_counts(dropna=False).reset_index()
        user_counts.columns = ['User', 'Verkauft']
        
        # Sortieren nach Anzahl Verkäufe (absteigend)
        user_counts = user_counts.sort_values('Verkauft', ascending=False)
        
        # GESAMT-Zeile am Ende hinzufügen
        total_sales = user_counts['Verkauft'].sum()
        gesamt_row = pd.DataFrame({
            'User': ['Gesamt'],
            'Verkauft': [total_sales]
        })
        
        # Kombiniere die Ergebnisse mit der Gesamtzeile
        sales_analysis = pd.concat([user_counts, gesamt_row], ignore_index=True)
        
        print("Verkaufsstatistik (Top 10):")
        print(sales_analysis.head(10))
        
        # Registerkarte hinzufügen
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sales_analysis.to_excel(writer, sheet_name='Verkäufe nach User', index=False)

            # Greifen auf das Worksheet-Objekt zu
            worksheet_user = writer.sheets['Verkäufe nach User']
            # Standardbreite ist oft ~8.43, wir setzen z.B. 25
            worksheet_user.column_dimensions['A'].width = 15
            worksheet_user.column_dimensions['B'].width = 15
        
        print(f"Verkaufsstatistik gespeichert in Registerkarte 'Verkäufe nach User'")
        print(f"Gesamtanzahl Verkäufe: {total_sales}")
        print(f"Anzahl verschiedener User: {len(user_counts)}")
        
    else:
        print("Nicht genügend Daten oder Spalten für Verkaufsstatistik")
        print(f"Verfügbare Zeilen: {len(df_final_grp) if 'df_final_grp' in locals() else 'df_final_grp nicht definiert'}, " +
              f"Verfügbare Spalten: {len(df_final_grp.columns) if 'df_final_grp' in locals() and len(df_final_grp) > 0 else 0}")
        
        # Leere Registerkarte erstellen
        empty_sales = pd.DataFrame({'Hinweis': ['Nicht genügend Daten für Verkaufsstatistik']})
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            empty_sales.to_excel(writer, sheet_name='Verkäufe nach User', index=False, header=False)
            
except Exception as e:
    print(f"Fehler bei der Verkaufsstatistik: {e}")
    # Kein traceback, da wir den Fehler bereits behandelt haben
    pass # Oder entfernen Sie den try/except Block, wenn Sie Fehler sehen wollen

# 15. User Regionen Tabelle erstellen
try:
    print(f"\n=== User Regionen Tabelle erstellen ===")
    
    # Erstelle DataFrame mit den User-Regionen Daten
    user_regionen_data = [
        ['AG0001', '5116', 'Mitte', 'Schiznach-Bad'],
        ['AG0127', '4310', 'Mitte', 'Rheinfelden'],
        ['AG0129', '4410', 'West', 'Liestal'],
        ['AG0133', '4058', 'West', 'Kleinbasel'],
        ['AG0136', '4133', 'West', 'Pratteln'],
        ['AG0137', '4132', 'West', 'Muttenz'],
        ['AG0140', '2555', 'West', 'Biel'],
        ['AG0150', '9470', 'Ost', 'Buchs'],
        ['AG0151', '9490', 'Ost', 'Vaduz'],
        ['AG0155', '4528', 'West', 'Solothurn'],
        ['AG0165', '7000', 'Ost', 'Chur'],
        ['AG0173', '6512', 'Mitte', 'Giubiasco'],
        ['AG0190', '7270', 'Ost', 'Davos'],
        ['AG0200', '8500', 'Ost', 'Frauenfeld'],
        ['AG0220', '1752', 'Romandie', 'Fribourg'],
        ['AG0227', '1630', 'Romandie', 'Bulle'],
        ['AG0232', '3280', 'West', 'Murten'],
        ['AG0244', '1197', 'Romandie', 'Nyon'],
        ['AG0248', '1260', 'Romandie', 'Nyon Champs Colin'],
        ['AG0260', '8280', 'Ost', 'Kreuzlingen'],
        ['AG0272', '8570', 'Ost', 'Weinfelden'],
        ['AG0289', '1219', 'Romandie', 'Genève'],
        ['AG0290', '3053', 'West', 'Occ. Münchenbuchsee'],
        ['AG0302', '1032', 'Romandie', 'Crissier'],
        ['AG0348', '1163', 'Romandie', 'Etoy'],
        ['AG0357', '9626', 'Mitte', 'Lugano'],
        ['AG0376', '6280', 'Mitte', 'Sursee'],
        ['AG0380', '6035', 'Mitte', 'Buchrain'],
        ['AG0399', '6215', 'Mitte', 'Beromünster'],
        ['AG0400', '6010', 'Mitte', 'Luzern'],
        ['ag0437', '4665', 'Mitte', 'Occ. Oftringen'],
        ['AG0450', '8200', 'Ost', 'Schaffhausen'],
        ['AG0460', '4657', 'West', 'Dulliken'],
        ['AG0464', '4665', 'Mitte', 'Oftringen'],
        ['AG0523', '9435', 'Ost', 'Heerbrugg'],
        ['AG0540', '5442', 'Mitte', 'Baden'],
        ['AG0570', '3604', 'West', 'Thun'],
        ['AG0620', '8406', 'Ost', 'Occ. Winterthur'],
        ['AG0650', '8008', 'Ost', 'Utoquai'],
        ['AG0655', '6340', 'Mitte', 'Sihlbrugg'],
        ['AG0659', '8051', 'Ost', 'Zürich'],
        ['AG0673', '8003', 'Mitte', 'Zürich Badenerstr.'],
        ['AG0674', '5430', 'Mitte', 'Wettingen'],
        ['AG0680', '8610', 'Ost', 'Uster'],
        ['AG0690', '8302', 'Ost', 'Kloten'],
        ['AG0698', '8600', 'Ost', 'Autowelt Audi'],
        ['AG0698SE', '8600', 'Ost', 'Autowelt Seat'],
        ['AG0698VW', '8600', 'Ost', 'Autowelt VW'],
        ['AG0710', '8048', 'Mitte', 'Letzigrund'],
        ['AG0739', '8810', 'Mitte', 'Horgen'],
        ['AG0760', '8846', 'Mitte', 'Jona'],
        ['AG0781', '8604', 'Ost', 'Volketswil'],
        ['AG0788', '8952', 'Mitte', 'Schlieren'],
        ['AG0790', '8184', 'Ost', 'Bülach'],
        ['AG3346', '3073', 'West', 'Gümligen'],
        ['AG3537', '5000', 'Mitte', 'Aarau'],
        ['AG3690', '6850', 'Mitte', 'Mendrisio'],
        ['AG6114', '1845', 'Romandie', 'Villeneuve'],
        ['AG6136', '1214', 'Romandie', 'Vernier'],
        ['AG6140', '2504', 'West', 'Längfeldweg'],
        ['AG6198', '8048', 'Mitte', 'Zürich'],
        ['AG6601', '3032', 'West', 'Emmen'],
        ['AG6772', '8907', 'Mitte', 'Wettswil'],
        ['AG6782', '8038', 'Mitte', 'Wollishofen']
    ]
    
    # Erstelle DataFrame
    df_user_regionen = pd.DataFrame(user_regionen_data, columns=['User', 'PLZ', 'Region', 'Stadort'])
    
    # Registerkarte hinzufügen
    with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_user_regionen.to_excel(writer, sheet_name='User Regionen', index=False)

        # Greifen auf das Worksheet-Objekt zu
        worksheet_amaguser = writer.sheets['User Regionen']
        # Standardbreite ist oft ~8.43, wir setzen z.B. 25
        worksheet_amaguser.column_dimensions['A'].width = 15
        worksheet_amaguser.column_dimensions['B'].width = 15
        worksheet_amaguser.column_dimensions['C'].width = 15
        worksheet_amaguser.column_dimensions['D'].width = 25
    
    print("Registerkarte 'User Regionen' hinzugefügt")
    print(f"Anzahl User: {len(df_user_regionen)}")
    
except Exception as e:
    print(f"Fehler beim Erstellen der User Regionen Tabelle: {e}")
    import traceback
    traceback.print_exc()


# 16. Kurzübersicht erstellen
try:
    print(f"\n=== Kurzübersicht_AMAG_{args.month} erstellen ===")
    
    # Überprüfen, ob alle benötigten Daten vorhanden sind
    if 'df_user_regionen' in locals() and 'df_final' in locals() and 'sales_analysis' in locals():
        
        # Kopiere die User-Regionen-Daten als Basis
        kurzuebersicht = df_user_regionen.copy()
        
        # Spalte "Beanstandungen" hinzufügen
        # Für jeden User (Spalte 0 in df_user_regionen) zähle die Einträge in df_final (Spalte 3)
        # Beginne ab der zweiten Zeile (ohne Header)
        beanstandungen_counts = df_final.iloc[1:, 3].value_counts()  
        
        # Erstelle ein Dictionary für schnelle Suche
        beanstandungen_dict = beanstandungen_counts.to_dict()
        
        # Füge die Spalte "Beanstandungen" hinzu
        kurzuebersicht['Beanstandungen'] = kurzuebersicht['User'].map(beanstandungen_dict).fillna(0).astype(int)
        
        # Spalte "Verkauft" aus sales_analysis hinzufügen
        # Erstelle ein Dictionary aus sales_analysis (User -> Verkauft)
        verkauft_dict = sales_analysis.set_index('User')['Verkauft'].to_dict()
        # Entferne die "Gesamt"-Zeile aus dem Dictionary, falls vorhanden
        verkauft_dict.pop('Gesamt', None)
        
        # Füge die Spalte "Verkauft" hinzu
        kurzuebersicht['Verkauft'] = kurzuebersicht['User'].map(verkauft_dict).fillna(0).astype(int)
        
        # Spalte "Beanstandungsquote(%)" hinzufügen
        # Vermeide Division durch Null
        kurzuebersicht['Beanstandungsquote(%)'] = np.where(
            kurzuebersicht['Verkauft'] > 0,
            (kurzuebersicht['Beanstandungen'] / kurzuebersicht['Verkauft'] * 100).round(2),
            0.00
        )
        
        # Gesamtzeile am Ende hinzufügen
        gesamt_row = pd.DataFrame({
            'User': ['Gesamt'],
            'PLZ': [''],
            'Region': [''],
            'Stadort': [''],
            'Beanstandungen': [kurzuebersicht['Beanstandungen'].sum()],
            'Verkauft': [kurzuebersicht['Verkauft'].sum()],
            'Beanstandungsquote(%)': [
                round(
                    (kurzuebersicht['Beanstandungen'].sum() / max(1, kurzuebersicht['Verkauft'].sum())) * 100,
                    2
                ) if kurzuebersicht['Verkauft'].sum() > 0 else 0.00
            ]
        })
        
        # Kombiniere die Daten mit der Gesamtzeile
        kurzuebersicht_final = pd.concat([kurzuebersicht, gesamt_row], ignore_index=True)
        
        # Füge die Registerkarte zur Excel-Datei hinzu
        sheet_name = f'Kurzübersicht_AMAG_{args.month}'
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            kurzuebersicht_final.to_excel(writer, sheet_name=sheet_name, index=False)

            # Greifen auf das Worksheet-Objekt zu
            worksheet_kurzübersicht = writer.sheets[sheet_name]
            # Standardbreite ist oft ~8.43, wir setzen z.B. 25
            worksheet_kurzübersicht.column_dimensions['A'].width = 10
            worksheet_kurzübersicht.column_dimensions['B'].width = 10
            worksheet_kurzübersicht.column_dimensions['C'].width = 15
            worksheet_kurzübersicht.column_dimensions['D'].width = 25
            worksheet_kurzübersicht.column_dimensions['E'].width = 15
            worksheet_kurzübersicht.column_dimensions['F'].width = 15
            worksheet_kurzübersicht.column_dimensions['G'].width = 25
        
        print(f"Registerkarte '{sheet_name}' hinzugefügt")
        print(f"Gesamt Beanstandungen: {kurzuebersicht['Beanstandungen'].sum()}")
        print(f"Gesamt Verkäufe: {kurzuebersicht['Verkauft'].sum()}")
        
    else:
        print("Nicht alle benötigten Daten sind verfügbar für die Kurzübersicht")
        # Erstelle eine leere Registerkarte mit einer Fehlermeldung
        error_data = pd.DataFrame({'Fehler': ['Benötigte Daten nicht verfügbar']})
        sheet_name = f'Kurzübersicht_AMAG_{args.month}'
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            error_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
except Exception as e:
    print(f"Fehler beim Erstellen der Kurzübersicht: {e}")
    import traceback
    traceback.print_exc()

# 17. Offene Fälle
try:
    print(f"\n=== Offene_Fälle_AMAG_{args.month} erstellen ===")
    
    # Überprüfen, ob alle benötigten Daten vorhanden sind
    if ('df_user_regionen' in locals() and 'erledigt_final' in locals() and 'offen_final' in locals() and len(offen_final) > 1):
        
        # Kopiere die User-Regionen-Daten als Basis
        offene_falle = df_user_regionen.copy()
        
        # Spalte "Beanstandungen" hinzufügen
        # Für jeden User (Spalte 0 in df_user_regionen) zähle die Einträge in df_final (Spalte 3)
        # Beginne ab der zweiten Zeile (ohne Header)
        abgeschlossene_counts = erledigt_final.iloc[1:, 3].value_counts()  
        
        # Erstelle ein Dictionary für schnelle Suche
        abgeschlossene_dict = abgeschlossene_counts.to_dict()
        
        # Füge die Spalte "Beanstandungen" hinzu
        offene_falle['Abgeschlossene Fälle'] = offene_falle['User'].map(abgeschlossene_dict).fillna(0).astype(int)
        
        # Spalte "Verkauft"  hinzufügen
        offene_counts = offen_final.iloc[1:, 3].value_counts()  
        offene_dict = offene_counts.to_dict()
        
        # Füge die Spalte "Verkauft" hinzu
        offene_falle['Offene Fälle'] = offene_falle['User'].map(offene_dict).fillna(0).astype(int)

        # Spalte "Begründung" hinzufügen ---
        # Extrahiereт die Datenzeilen aus offen_final (ohne Header)
        offen_data_rows = offen_final.iloc[1:]
        
        # Gruppiere nach User (Spalte 3 in offen_final) und sammle eindeutige Begründungen (Spalte 9 in offen_final)
        # Verwende dropna=False, um auch leere Begründungen zu berücksichtigen, falls nötig
        begruendungen_grouped = offen_data_rows.groupby(offen_data_rows.iloc[:, 3])[offen_data_rows.columns[9]].apply(
            lambda x: '  \n \n'.join(sorted(x.dropna().astype(str).unique())) # Verwenden '  ' (zwei Leerzeichen) als Trenner
        )

        # Erstellen ein Dictionary aus der Gruppierung
        begruendungen_dict = begruendungen_grouped.to_dict()

        # Füge die Spalte "Begründung" hinzu
        offene_falle['Begründung'] = offene_falle['User'].map(begruendungen_dict).fillna('')

        # Leere Spalten hinzufügen 
        offene_falle['Hängig bei CA'] = ''   # Leere Spalte
        offene_falle['Hängig bei AMAG'] = '' # Leere Spalte

        # Gesamtzeile am Ende hinzufügen
        gesamt_row = pd.DataFrame({
            'User': ['Gesamt'],
            'PLZ': [''],
            'Region': [''],
            'Stadort': [''],
            'Abgeschlossene Fälle': [offene_falle['Abgeschlossene Fälle'].sum()],
            'Offene Fälle': [offene_falle['Offene Fälle'].sum()],
            'Begründung': [''],
            'Hängig bei CA': [''],
            'Hängig bei AMAG': ['']  
        })
        
        # Kombiniere die Daten mit der Gesamtzeile
        offene_final = pd.concat([offene_falle, gesamt_row], ignore_index=True)
        
        # Füge die Registerkarte zur Excel-Datei hinzu
        sheet_name = f'Offene_Fälle_AMAG_{args.month}'
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            offene_final.to_excel(writer, sheet_name=sheet_name, index=False)

            from openpyxl import load_workbook

            worksheet_offen = writer.sheets[sheet_name] 
            from openpyxl.utils import get_column_letter
            spalte_begründung_buchstabe = get_column_letter(7)

            # Iteriereт durch alle Zeilen mit Daten in dieser Spalte (beginnend mit Zeile 2, da Zeile 1 der Header ist)
            for reihe in range(2, len(offene_final) + 2): # +2 weil: 1-basiert + Header-Zeile
                zelle = worksheet_offen[f'{spalte_begründung_buchstabe}{reihe}']
                # Aktiviere den automatischen Zeilenumbruch für diese Zelle
                zelle.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

            # Greifen auf das Worksheet-Objekt zu
            worksheet_kurzübersicht = writer.sheets[sheet_name]
            # Standardbreite ist oft ~8.43, wir setzen z.B. 25
            worksheet_offen.column_dimensions['A'].width = 10
            worksheet_offen.column_dimensions['B'].width = 10
            worksheet_offen.column_dimensions['C'].width = 15
            worksheet_offen.column_dimensions['D'].width = 25
            worksheet_offen.column_dimensions['E'].width = 25
            worksheet_offen.column_dimensions['F'].width = 25
            worksheet_offen.column_dimensions['G'].width = 45
            worksheet_offen.column_dimensions['H'].width = 20 
            worksheet_offen.column_dimensions['I'].width = 20
        
        print(f"Registerkarte '{sheet_name}' hinzugefügt")
        
    else:
        print("Nicht alle benötigten Daten sind verfügbar für die Offen Fälle")
        # Erstelle eine leere Registerkarte mit einer Fehlermeldung
        error_data = pd.DataFrame({'Fehler': ['Benötigte Daten nicht verfügbar']})
        sheet_name = f'Offene_Fälle_AMAG_{args.month}'
        with pd.ExcelWriter(result_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            error_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
except Exception as e:
    print(f"Fehler beim Erstellen der Offene Fälle: {e}")
    import traceback
    traceback.print_exc()


print(f"\nFertig! Ergebnis gespeichert in: {result_filename}")
print("Verfügbare Registerkarten:")
print("1. 'Alle' - Alle gefilterten Daten")
print("2. 'Erledigt' - Nur erledigte Beanstandungen")
print("3. 'Offen' - Nur offene Beanstandungen")
print("4. 'Hauptthema Analyse Ergebnis' - Statistische Auswertung")
print("5. 'Pivot Einsteller Hauptthema' - Kreuztabelle Einsteller x Hauptthema")
print("6. 'Gruppenreporting'")
print("7. 'Verkäufe nach User'")
print("8. 'User Regionen'")
print("9. 'Kurzübersicht'")
print("10. 'Offene Fälle'")